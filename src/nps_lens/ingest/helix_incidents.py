from __future__ import annotations

from hashlib import sha1
from typing import List, Optional, Union

import pandas as pd

from nps_lens.ingest.base import IngestResult, ValidationIssue, require_columns, standardize_columns

# Canonical context columns used by the app.
# We require N1 (channel) to be present. Company/N2 are optional because some
# Helix extracts are already pre-filtered or omit those fields.
HELIX_REQUIRED = [
    "BBVA_SourceServiceN1",
]


def dataset_id_for(path: str, service_origin: str, service_origin_n1: str) -> str:
    h = sha1(f"{path}|{service_origin}|{service_origin_n1}|helix".encode("utf-8")).hexdigest()[:10]
    return f"helix_incidents:{service_origin}:{service_origin_n1}:{h}"


def _split_csvish(value: object) -> List[str]:
    if value is None or (isinstance(value, float) and pd.isna(value)):
        return []
    s = str(value).strip()
    if not s:
        return []
    return [p.strip() for p in s.split(",") if p.strip()]


def _detect_fecha_column(df: pd.DataFrame) -> Optional[str]:
    """Best-effort date column detection for Helix incident exports.

    We normalize to a canonical `Fecha` used by storage partitioning.
    """
    candidates = [
        "Fecha",
        "Fecha apertura",
        "Fecha Apertura",
        "Fecha creación",
        "Fecha creacion",
        "Submit Date",
        "SubmitDate",
        "Submitted Date",
        "SubmittedDate",
        "CreatedDate",
        "Created Date",
        "Open Date",
        "Date",
    ]
    for c in candidates:
        if c in df.columns:
            return c
    # fallback: first column that contains 'fecha' or 'date'
    for c in df.columns:
        lc = str(c).lower()
        if "fecha" in lc or "date" in lc:
            return c
    return None


def _parse_helix_datetime(series: pd.Series) -> pd.Series:
    """Parse Helix datetime values robustly.

    Helix exports (and API/log-derived extracts) commonly encode timestamps as
    Unix epoch **milliseconds** (e.g. 1767576293000). Pandas' default
    to_datetime() can misinterpret these depending on dtype.

    Strategy:
      1) Try regular to_datetime (handles ISO strings, Excel datetimes, etc.).
      2) If most values are NaT and the series looks numeric, interpret as:
         - milliseconds if magnitude ~ 1e12 or higher
         - seconds if magnitude ~ 1e9

    Output is timezone-naive to keep analysis consistent across the app.
    """

    s = series.copy()

    def _epoch_to_dt(num: pd.Series) -> pd.Series:
        """Convert numeric epoch series to datetime (robust to mixed units).

        Helix extracts are usually **epoch milliseconds**, but in practice we
        also see seconds, microseconds or nanoseconds (e.g. logs / API joins).
        Passing a nanosecond value with unit='ms' triggers overflow.

        Strategy:
          - Coerce to numeric.
          - If the column is "mostly numeric", detect unit by magnitude.
          - Handle mixed magnitudes by converting subsets with different units.

        Output is timezone-naive.
        """

        n = pd.to_numeric(num, errors="coerce")
        if len(n) == 0 or float(n.notna().mean()) < 0.6:
            return pd.to_datetime(n, errors="coerce")

        abs_n = n.abs()
        # Heuristic thresholds (order matters):
        #   ns ~ 1e18, us ~ 1e15, ms ~ 1e12, s ~ 1e9
        mask_ns = abs_n >= 1e17
        mask_us = (abs_n >= 1e14) & ~mask_ns
        mask_ms = (abs_n >= 1e11) & ~(mask_ns | mask_us)
        mask_s = (abs_n >= 1e9) & ~(mask_ns | mask_us | mask_ms)

        out = pd.Series(pd.NaT, index=n.index)

        def _convert(mask: pd.Series, unit: str) -> None:
            if not bool(mask.any()):
                return
            out.loc[mask] = pd.to_datetime(n.loc[mask], unit=unit, utc=True, errors="coerce").dt.tz_localize(None)

        _convert(mask_ns, "ns")
        _convert(mask_us, "us")
        _convert(mask_ms, "ms")
        _convert(mask_s, "s")

        # If some numeric values are too small for epoch heuristics, try generic.
        mask_rest = n.notna() & out.isna()
        if bool(mask_rest.any()):
            out.loc[mask_rest] = pd.to_datetime(n.loc[mask_rest], errors="coerce")

        return out

    # 1) If already numeric -> treat as epoch first (avoid ns default)
    if pd.api.types.is_numeric_dtype(s):
        return _epoch_to_dt(s)

    # 2) If object dtype -> try to coerce to numeric epoch (handles thousands separators)
    try:
        cleaned = s.astype("string").str.replace(r"[^0-9\\-]", "", regex=True)
        num = pd.to_numeric(cleaned, errors="coerce")
        if len(num) and float(num.notna().mean()) >= 0.6:
            return _epoch_to_dt(num)
    except Exception:
        pass

    # 3) Fallback: general parser for ISO/excel date strings
    return pd.to_datetime(s, errors="coerce")


def _looks_like_datetime_col(col: str) -> bool:
    lc = str(col).lower()
    return (
        "fecha" in lc
        or "date" in lc
        or "datetime" in lc
        or "timestamp" in lc
        or "datt" in lc
        or lc.endswith("_date")
        or lc.endswith("_datetime")
    )


def _auto_parse_epoch_datetime_columns(d: pd.DataFrame, issues: List[ValidationIssue]) -> pd.DataFrame:
    """Convert epoch-encoded datetime columns to pandas datetime.

    Many Helix exports include multiple timestamp fields stored as Unix epoch
    milliseconds (or seconds). To make downstream slicing and debugging easy,
    we convert *all* date-like columns (by name heuristic) to datetime when they
    look numeric.

    We preserve non-date columns and avoid coercing small numeric fields.
    """

    out = d.copy()
    converted: List[str] = []
    for c in list(out.columns):
        if c == "Fecha":
            continue
        if not _looks_like_datetime_col(str(c)):
            continue

        dt = _parse_helix_datetime(out[c])
        if len(dt) and float(dt.notna().mean()) >= 0.6:
            out[c] = dt
            converted.append(str(c))

    if converted:
        issues.append(
            ValidationIssue(
                level="INFO",
                message=(
                    "Columnas de fecha detectadas y convertidas desde epoch/strings a datetime: "
                    + ", ".join(converted[:20])
                    + (" ..." if len(converted) > 20 else "")
                ),
            )
        )
    return out


def read_helix_incidents_excel(
    path: str,
    service_origin: str,
    service_origin_n1: str,
    service_origin_n2: str,
    sheet_name: Optional[Union[str, int]] = None,
) -> IngestResult:
    """Read + filter Helix incidents Excel by selected context.

    Contract (strict filtering):
      - Always filter by:
          service_origin == BBVA_SourceServiceCompany
          service_origin_n1 == BBVA_SourceServiceN1
      - Only if the selected context has service_origin_n2 tokens (non-empty),
        then ALSO filter by *strict token-set equality*: keep rows whose
        BBVA_SourceServiceN2 token-set equals the selected token-set.

    If after filtering there are no rows, return empty df (ingestion is not performed).
    """

    # Prefer Helix_Raw / Helix raw sheet as source of truth (not "Issues oficial").
    if sheet_name is None:
        try:
            import openpyxl  # type: ignore
            wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
            sheetnames = list(wb.sheetnames)
        except Exception:
            sheetnames = []
        candidates = ["Helix_Raw", "Helix raw", "Helix Raw", "helix_raw", "helix raw"]
        picked = None
        lower_map = {s.lower(): s for s in sheetnames}
        for c in candidates:
            if c.lower() in lower_map:
                picked = lower_map[c.lower()]
                break
        sn: Union[str, int] = picked if picked is not None else 0
    else:
        sn = sheet_name
    df = pd.read_excel(path, sheet_name=sn, engine="openpyxl")
    if isinstance(df, dict):
        df = list(df.values())[0]

    # Canonicalize / robust column names (tolerate minor variants)
    df = standardize_columns(
        df,
        mapping={
            "BBVA_SourceServiceCompany": "BBVA_SourceServiceCompany",
            "BBVA Source Service Company": "BBVA_SourceServiceCompany",
            "SourceServiceCompany": "BBVA_SourceServiceCompany",
            "Servicio Origen - BU/UG": "BBVA_SourceServiceCompany",
            "BBVA_SourceServiceN1": "BBVA_SourceServiceN1",
            "BBVA Source Service N1": "BBVA_SourceServiceN1",
            "SourceServiceN1": "BBVA_SourceServiceN1",
            "Servicio Origen - Servicio N1": "BBVA_SourceServiceN1",
            "BBVA_SourceServiceN2": "BBVA_SourceServiceN2",
            "BBVA Source Service N2": "BBVA_SourceServiceN2",
            "SourceServiceN2": "BBVA_SourceServiceN2",
            "Servicio Origen - Servicio N2": "BBVA_SourceServiceN2",
        },
    )

    issues: List[ValidationIssue] = []
    issues.extend(require_columns(df, HELIX_REQUIRED))

    if any(i.level == "ERROR" for i in issues):
        return IngestResult(df=df, issues=issues, dataset_id=dataset_id_for(path, service_origin, service_origin_n1))

    # Ensure optional canonical context columns exist.
    if "BBVA_SourceServiceCompany" not in df.columns:
        df["BBVA_SourceServiceCompany"] = str(service_origin)
        issues.append(
            ValidationIssue(
                level="WARN",
                message=(
                    "Columna de contexto 'BBVA_SourceServiceCompany' no encontrada en el fichero Helix. "
                    "Se asume que el fichero ya viene filtrado para el service_origin seleccionado."
                ),
            )
        )
    if "BBVA_SourceServiceN2" not in df.columns:
        df["BBVA_SourceServiceN2"] = ""

    # Normalize N2 column to stable CSV-ish string
    d = df.copy()
    d["BBVA_SourceServiceCompany"] = d["BBVA_SourceServiceCompany"].astype(str)
    d["BBVA_SourceServiceN1"] = d["BBVA_SourceServiceN1"].astype(str)
    d["BBVA_SourceServiceN2"] = d["BBVA_SourceServiceN2"].apply(lambda v: ", ".join(_split_csvish(v)))

    # Context filters
    # Company filter is best-effort: if the column was missing we filled it with the selected origin.
    if "BBVA_SourceServiceCompany" in d.columns:
        before = len(d)
        d = d.loc[d["BBVA_SourceServiceCompany"].astype(str) == str(service_origin)]
        dropped = before - len(d)
        if dropped:
            issues.append(
                ValidationIssue(
                    level="INFO",
                    message=f"Filtradas {dropped} filas fuera de BBVA_SourceServiceCompany={service_origin}.",
                )
            )

    before = len(d)
    d = d.loc[d["BBVA_SourceServiceN1"].astype(str) == str(service_origin_n1)]
    dropped = before - len(d)
    if dropped:
        issues.append(
            ValidationIssue(
                level="INFO",
                message=f"Filtradas {dropped} filas fuera de BBVA_SourceServiceN1={service_origin_n1}.",
            )
        )

    # Optional N2 filter ONLY when selected context has tokens.
    # Semantics: strict token-set equality (order-insensitive).
    sel_n2 = [v.strip() for v in (service_origin_n2 or "").split(",") if v.strip()]
    if sel_n2:
        sel = {v.strip() for v in sel_n2 if v.strip()}

        def _row_matches_exact(v: object) -> bool:
            toks = {p.strip() for p in str(v or "").split(",") if p.strip()}
            return toks == sel

        before = len(d)
        d = d.loc[d["BBVA_SourceServiceN2"].apply(_row_matches_exact)]
        dropped = before - len(d)
        if dropped:
            issues.append(
                ValidationIssue(
                    level="INFO",
                    message=(
                        f"Filtradas {dropped} filas fuera de BBVA_SourceServiceN2 == {{{', '.join(sorted(sel))}}}."
                    ),
                )
            )

    # If empty after filtering, signal to caller (no persistence)
    if d.empty:
        issues.append(
            ValidationIssue(
                level="WARN",
                message=(
                    "No hay registros para el contexto seleccionado. "
                    "La ingesta se omite para evitar mezclar contextos."
                ),
            )
        )
        return IngestResult(df=d, issues=issues, dataset_id=dataset_id_for(path, service_origin, service_origin_n1))

    # Attach selected context columns for downstream joins
    d["service_origin"] = str(service_origin)
    d["service_origin_n1"] = str(service_origin_n1)
    d["service_origin_n2_selected"] = ", ".join(sel_n2)

    # Canonical Fecha (best-effort)
    fecha_col = _detect_fecha_column(d)
    if fecha_col is not None:
        d["Fecha"] = _parse_helix_datetime(d[fecha_col])
        bad = int(d["Fecha"].isna().sum())
        if bad:
            issues.append(ValidationIssue(level="WARN", message=f"{bad} filas con Fecha inválida (columna '{fecha_col}')"))
    else:
        d["Fecha"] = pd.NaT
        issues.append(
            ValidationIssue(
                level="WARN",
                message="No se detectó columna de fecha. Se guardará Fecha=NaT (sin particionado temporal).",
            )
        )

    # Convert other date-like columns (many come as epoch ms) for easier inspection and consistent slicing
    d = _auto_parse_epoch_datetime_columns(d, issues)

    return IngestResult(df=d, issues=issues, dataset_id=dataset_id_for(path, service_origin, service_origin_n1))